"""ExpensIQ — Bank statement parser for Rural Kutxa / Ruralvía CSV & Excel."""

import csv
import hashlib
import io
import logging
import re
from datetime import date, datetime
from typing import Optional

logger = logging.getLogger("expensiq.bank_parser")

# Column name variants we recognize (case-insensitive, stripped)
DATE_COLS = {"fecha", "fecha operacion", "fecha operación", "f. operacion", "f.operacion", "date"}
VALUE_DATE_COLS = {"fecha valor", "f. valor", "f.valor", "value date"}
CONCEPT_COLS = {"concepto", "descripcion", "descripción", "comercio", "concept", "merchant", "detalle"}
AMOUNT_COLS = {"importe", "cantidad", "amount", "monto", "cargo", "importe (eur)", "importe(eur)"}
BALANCE_COLS = {"saldo", "balance", "saldo disponible"}
REFERENCE_COLS = {"referencia", "reference", "ref", "nº operacion", "nº operación", "num. operacion"}


def _normalize_col(name: str) -> str:
    """Strip whitespace, BOM, and lowercase."""
    return name.strip().strip("\ufeff").lower()


def _parse_spanish_number(raw: str) -> Optional[float]:
    """Parse '1.234,56' or '-1.234,56' or '1234.56' to float."""
    if not raw:
        return None
    s = raw.strip().replace(" ", "")
    # Detect Spanish format: has comma as decimal separator
    if "," in s and "." in s:
        # 1.234,56 → 1234.56
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        # 234,56 → 234.56
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(raw: str) -> Optional[date]:
    """Parse dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd."""
    if not raw:
        return None
    s = raw.strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _detect_separator(sample: str) -> str:
    """Detect CSV separator: ; or , or \\t."""
    counts = {";": sample.count(";"), ",": sample.count(","), "\t": sample.count("\t")}
    return max(counts, key=counts.get)  # type: ignore[arg-type]


def _make_external_id(ref: Optional[str], txn_date: Optional[date], merchant: Optional[str], amount: Optional[float]) -> str:
    """Generate a deterministic external_id from transaction fields."""
    if ref and ref.strip():
        return f"RK-{ref.strip()}"
    # Fallback: hash of date+merchant+amount
    raw = f"{txn_date or ''}-{merchant or ''}-{amount or ''}"
    h = hashlib.md5(raw.encode()).hexdigest()[:12]
    return f"RK-{h}"


def _map_columns(headers: list[str]) -> dict[str, int]:
    """Map semantic column names to their index."""
    mapping: dict[str, int] = {}
    for i, h in enumerate(headers):
        norm = _normalize_col(h)
        if norm in DATE_COLS:
            mapping.setdefault("date", i)
        elif norm in VALUE_DATE_COLS:
            mapping.setdefault("value_date", i)
        elif norm in CONCEPT_COLS:
            mapping.setdefault("merchant", i)
        elif norm in AMOUNT_COLS:
            mapping.setdefault("amount", i)
        elif norm in BALANCE_COLS:
            mapping.setdefault("balance", i)
        elif norm in REFERENCE_COLS:
            mapping.setdefault("reference", i)
    return mapping


class BankStatementParser:
    """Parses Rural Kutxa / Ruralvía bank statements in CSV or Excel format."""

    def parse(self, file_content: bytes, filename: str) -> list[dict]:
        """Parse file and return list of transaction dicts."""
        lower = filename.lower()
        if lower.endswith((".xlsx", ".xls")):
            return self._parse_excel(file_content)
        else:
            return self._parse_csv(file_content)

    def _parse_csv(self, content: bytes) -> list[dict]:
        """Parse CSV with auto-detection of encoding, separator, and columns."""
        # Try encodings
        text = None
        for enc in ("utf-8-sig", "utf-8", "latin-1", "iso-8859-1", "cp1252"):
            try:
                text = content.decode(enc)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        if text is None:
            text = content.decode("utf-8", errors="replace")

        # Skip leading empty lines or bank header lines
        lines = text.strip().splitlines()
        start = 0
        for i, line in enumerate(lines):
            stripped = line.strip()
            if not stripped:
                continue
            # Check if this looks like a header (has known column names)
            norm = _normalize_col(stripped)
            if any(kw in norm for kw in ("fecha", "date", "importe", "amount", "concepto")):
                start = i
                break
            # If it has enough separators, it might be data — check next line too
            if i > 5:
                start = 0
                break

        if start > 0:
            lines = lines[start:]

        text_clean = "\n".join(lines)
        sep = _detect_separator(text_clean)

        reader = csv.reader(io.StringIO(text_clean), delimiter=sep)
        rows_list = list(reader)
        if not rows_list:
            return []

        col_map = _map_columns(rows_list[0])
        if "amount" not in col_map:
            # Try second row as header
            if len(rows_list) > 1:
                col_map = _map_columns(rows_list[1])
                rows_list = rows_list[1:]
            if "amount" not in col_map:
                logger.warning("Could not detect amount column in CSV")
                return []

        results = []
        for row in rows_list[1:]:  # Skip header
            if not row or len(row) <= max(col_map.values()):
                continue
            # Skip empty rows
            amount_raw = row[col_map["amount"]].strip() if "amount" in col_map else ""
            if not amount_raw:
                continue

            amount = _parse_spanish_number(amount_raw)
            if amount is None:
                continue

            txn_date = _parse_date(row[col_map["date"]]) if "date" in col_map else None
            merchant = row[col_map["merchant"]].strip() if "merchant" in col_map else None
            reference = row[col_map["reference"]].strip() if "reference" in col_map else None

            results.append({
                "date": txn_date.isoformat() if txn_date else None,
                "merchant": merchant or None,
                "amount": abs(amount),  # Store as positive (expenses are charges)
                "reference": reference or None,
                "external_id": _make_external_id(reference, txn_date, merchant, amount),
            })

        logger.info("Parsed %d transactions from CSV", len(results))
        return results

    def _parse_excel(self, content: bytes) -> list[dict]:
        """Parse Excel file using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl not installed — cannot parse Excel files")
            return []

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Find header row (first row with recognizable column names)
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            str_row = [str(c or "").strip() for c in row]
            col_map = _map_columns(str_row)
            if "amount" in col_map:
                header_idx = i
                break
        else:
            logger.warning("Could not detect header row in Excel")
            return []

        headers = [str(c or "").strip() for c in rows[header_idx]]
        col_map = _map_columns(headers)

        results = []
        for row in rows[header_idx + 1:]:
            cells = list(row)
            if len(cells) <= max(col_map.values()):
                continue

            # Amount
            amount_cell = cells[col_map["amount"]]
            if amount_cell is None:
                continue
            if isinstance(amount_cell, (int, float)):
                amount = float(amount_cell)
            else:
                amount = _parse_spanish_number(str(amount_cell))
            if amount is None:
                continue

            # Date
            date_cell = cells[col_map["date"]] if "date" in col_map else None
            if isinstance(date_cell, datetime):
                txn_date = date_cell.date()
            elif isinstance(date_cell, date):
                txn_date = date_cell
            else:
                txn_date = _parse_date(str(date_cell or ""))

            # Merchant
            merchant = str(cells[col_map["merchant"]]).strip() if "merchant" in col_map and cells[col_map["merchant"]] else None

            # Reference
            reference = str(cells[col_map["reference"]]).strip() if "reference" in col_map and cells[col_map["reference"]] else None

            results.append({
                "date": txn_date.isoformat() if txn_date else None,
                "merchant": merchant,
                "amount": abs(amount),
                "reference": reference,
                "external_id": _make_external_id(reference, txn_date, merchant, amount),
            })

        wb.close()
        logger.info("Parsed %d transactions from Excel", len(results))
        return results
