"""ExpensIQ — Receipt routes."""

import csv
import io
import logging
import uuid
from datetime import date as DateType, datetime
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, Header, HTTPException, Response, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.auth import get_current_user_optional
from app.core.config import settings
from app.db.session import get_db
from app.models.models import BankTransaction, Employee, Match, Project, Receipt
from app.schemas.schemas import ApproveRejectResult, ReceiptMatchOut, ReceiptOut, ReceiptUpdate, ReconcileSingleResult

logger = logging.getLogger("expensiq.receipts")

router = APIRouter()

ALLOWED_TYPES = {
    "image/jpeg", "image/png", "image/webp", "image/gif", "application/pdf",
}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


@router.get("", response_model=list[ReceiptOut])
def list_receipts(
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    category: Optional[str] = None,
    project_id: Optional[str] = None,
    date_from: Optional[DateType] = None,
    date_to: Optional[DateType] = None,
    search: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_order: Optional[str] = "desc",
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    from sqlalchemy.orm import joinedload
    query = db.query(Receipt).options(
        joinedload(Receipt.employee),
        joinedload(Receipt.approver),
        joinedload(Receipt.project),
    )
    if status:
        query = query.filter(Receipt.status == status)
    if employee_id:
        query = query.filter(Receipt.employee_id == employee_id)
    if category:
        query = query.filter(Receipt.category == category)
    if project_id:
        query = query.filter(Receipt.project_id == project_id)
    if date_from:
        query = query.filter(Receipt.date >= date_from)
    if date_to:
        query = query.filter(Receipt.date <= date_to)
    if search:
        query = query.filter(Receipt.merchant.ilike(f"%{search}%"))

    # Sorting
    sort_column = {
        "date": Receipt.date,
        "amount": Receipt.amount,
        "merchant": Receipt.merchant,
        "category": Receipt.category,
        "status": Receipt.status,
    }.get(sort_by, Receipt.upload_timestamp)

    if sort_order == "asc":
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())

    return query.offset(offset).limit(limit).all()


@router.get("/export/csv")
def export_receipts_csv(
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    category: Optional[str] = None,
    project_id: Optional[str] = None,
    date_from: Optional[DateType] = None,
    date_to: Optional[DateType] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
):
    from sqlalchemy.orm import joinedload
    query = db.query(Receipt).options(joinedload(Receipt.employee), joinedload(Receipt.project))
    if status:
        query = query.filter(Receipt.status == status)
    if employee_id:
        query = query.filter(Receipt.employee_id == employee_id)
    if category:
        query = query.filter(Receipt.category == category)
    if project_id:
        query = query.filter(Receipt.project_id == project_id)
    if date_from:
        query = query.filter(Receipt.date >= date_from)
    if date_to:
        query = query.filter(Receipt.date <= date_to)
    if search:
        query = query.filter(Receipt.merchant.ilike(f"%{search}%"))
    receipts = query.order_by(Receipt.upload_timestamp.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "id", "empleado", "obra", "merchant", "fecha", "importe", "moneda",
        "base_imponible", "tipo_iva", "cuota_iva", "tax_legacy",
        "categoria", "estado", "confianza_ocr", "fecha_subida", "notas",
    ])
    for r in receipts:
        writer.writerow([
            str(r.id),
            r.employee.name if r.employee else "",
            r.project.code if r.project else "",
            r.merchant or "",
            str(r.date) if r.date else "",
            r.amount or "",
            r.currency,
            r.tax_base or "",
            r.tax_rate or "",
            r.tax_amount or "",
            r.tax or "",
            r.category,
            r.status,
            round(float(r.ocr_confidence), 2) if r.ocr_confidence else "",
            str(r.upload_timestamp) if r.upload_timestamp else "",
            r.notes or "",
        ])
    output.seek(0)
    filename = f"gastos_{datetime.now().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/upload", response_model=ReceiptOut, status_code=201)
async def upload_receipt(
    background_tasks: BackgroundTasks,
    employee_id: str = Form(...),
    file: UploadFile = File(...),
    project_id: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    # Validate employee exists
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Validate file type
    if file.content_type not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail=f"File type not allowed: {file.content_type}")

    # Read file and check size
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File exceeds 10 MB limit")

    # Store file
    receipt_id = uuid.uuid4()
    object_key = f"receipts/{receipt_id}/{file.filename}"
    image_url = None

    try:
        from app.services.storage import storage_service
        storage_service.upload(object_key, content, file.content_type)
        image_url = f"{settings.S3_ENDPOINT}/{settings.S3_BUCKET}/{object_key}"
    except Exception as e:
        logger.warning("Storage upload failed, continuing without image_url: %s", e)

    # Validate project if provided
    resolved_project_id = None
    if project_id:
        project = db.query(Project).filter(Project.id == project_id, Project.active == True).first()
        if not project:
            raise HTTPException(status_code=404, detail="Obra no encontrada o inactiva")
        resolved_project_id = project.id

    # Create receipt record
    receipt = Receipt(
        id=receipt_id,
        employee_id=employee_id,
        image_url=image_url,
        status="processing",
        project_id=resolved_project_id,
    )
    db.add(receipt)
    db.commit()
    db.refresh(receipt)

    # Trigger background OCR processing
    background_tasks.add_task(_process_receipt_ocr, str(receipt_id), content, file.filename)

    return receipt


@router.get("/{receipt_id}", response_model=ReceiptOut)
def get_receipt(receipt_id: str, db: Session = Depends(get_db)):
    from sqlalchemy.orm import joinedload
    receipt = (
        db.query(Receipt)
        .options(
            joinedload(Receipt.employee),
            joinedload(Receipt.approver),
            joinedload(Receipt.project),
        )
        .filter(Receipt.id == receipt_id)
        .first()
    )
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")
    return receipt


@router.get("/{receipt_id}/image")
def get_receipt_image(receipt_id: str, db: Session = Depends(get_db)):
    """Proxy receipt image from MinIO so the browser can display it."""
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt or not receipt.image_url:
        raise HTTPException(status_code=404, detail="Image not found")

    # Extract the S3 object key from the stored URL
    # URL format: http://minio:9000/bucket/receipts/uuid/filename
    try:
        from app.services.storage import storage_service
        url_path = receipt.image_url.split(f"/{settings.S3_BUCKET}/", 1)[1]
        data = storage_service.download(url_path)
    except Exception:
        # Fallback: try serving from static files
        if receipt.image_url.startswith("/static/"):
            return Response(status_code=302, headers={"Location": receipt.image_url})
        raise HTTPException(status_code=404, detail="Could not retrieve image")

    # Guess content type from filename
    ext = url_path.rsplit(".", 1)[-1].lower() if "." in url_path else "jpeg"
    ct = {"jpeg": "image/jpeg", "jpg": "image/jpeg", "png": "image/png", "webp": "image/webp", "gif": "image/gif", "pdf": "application/pdf"}.get(ext, "image/jpeg")

    return Response(content=data, media_type=ct)


@router.get("/{receipt_id}/matches", response_model=list[ReceiptMatchOut])
def get_receipt_matches(receipt_id: str, db: Session = Depends(get_db)):
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")

    matches = (
        db.query(Match, BankTransaction)
        .join(BankTransaction, Match.transaction_id == BankTransaction.id)
        .filter(Match.receipt_id == receipt_id)
        .all()
    )

    return [
        ReceiptMatchOut(
            match_id=m.id,
            transaction_id=t.id,
            confidence=float(m.confidence) if m.confidence else None,
            match_method=m.match_method,
            transaction_date=t.date,
            transaction_merchant=t.merchant,
            transaction_amount=float(t.amount),
            transaction_currency=t.currency,
        )
        for m, t in matches
    ]


@router.post("/{receipt_id}/reconcile", response_model=ReconcileSingleResult)
def reconcile_single(receipt_id: str, db: Session = Depends(get_db)):
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")

    from app.services.reconciliation import reconciliation_engine
    matches_created = reconciliation_engine.reconcile_receipt(db, receipt)

    return ReconcileSingleResult(status=receipt.status, matches_created=matches_created)


@router.patch("/{receipt_id}", response_model=ReceiptOut)
def update_receipt(receipt_id: str, data: ReceiptUpdate, db: Session = Depends(get_db)):
    from sqlalchemy.orm import joinedload
    receipt = (
        db.query(Receipt)
        .options(joinedload(Receipt.employee), joinedload(Receipt.approver), joinedload(Receipt.project))
        .filter(Receipt.id == receipt_id)
        .first()
    )
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")
    updates = data.model_dump(exclude_unset=True)
    # project_id needs special handling: string → UUID or None
    if "project_id" in updates:
        pid = updates.pop("project_id")
        if pid is None or pid == "":
            receipt.project_id = None
        else:
            project = db.query(Project).filter(Project.id == pid, Project.active == True).first()
            if not project:
                raise HTTPException(status_code=404, detail="Obra no encontrada o inactiva")
            receipt.project_id = project.id
    for field, value in updates.items():
        setattr(receipt, field, value)
    db.commit()
    db.refresh(receipt)
    return receipt


# Fallback approval thresholds in EUR — used when the settings service has no
# DB session available. Sprint 1 fase B: los valores reales viven en tabla
# `settings` (keys approval.threshold_auto / approval.threshold_manager /
# approval.auto_enabled) y son editables desde /settings.
APPROVAL_THRESHOLD_AUTO = 100.0
APPROVAL_THRESHOLD_MANAGER = 500.0


def _calculate_approval_level(
    amount: float | None,
    db: Optional[Session] = None,
) -> str:
    """Determine approval level based on amount (3 tiers).

    If `db` is provided, thresholds are loaded from the settings table (cached
    briefly in-process). If `db` is None, falls back to module-level defaults
    so callers without a session (tests, scripts) keep working.
    """
    from app.services.settings_service import get_approval_thresholds

    threshold_auto = APPROVAL_THRESHOLD_AUTO
    threshold_manager = APPROVAL_THRESHOLD_MANAGER
    auto_enabled = True
    if db is not None:
        thresholds = get_approval_thresholds(db)
        threshold_auto = float(thresholds["threshold_auto"] or APPROVAL_THRESHOLD_AUTO)
        threshold_manager = float(thresholds["threshold_manager"] or APPROVAL_THRESHOLD_MANAGER)
        auto_enabled = bool(thresholds["auto_enabled"])

    # If auto-approval disabled globally, bump everything out of "auto" tier
    # so no receipt self-approves regardless of amount.
    if not auto_enabled:
        if amount is None or amount < threshold_manager:
            return "manager"
        return "director"

    if amount is None or amount < threshold_auto:
        return "auto"
    if amount < threshold_manager:
        return "manager"
    return "director"


def _approval_reason(amount: float | None, level: str, db: Optional[Session] = None) -> str | None:
    """Human-readable reason why this receipt needs approval (null if auto)."""
    if level == "auto" or amount is None:
        return None
    threshold_auto = APPROVAL_THRESHOLD_AUTO
    threshold_manager = APPROVAL_THRESHOLD_MANAGER
    if db is not None:
        from app.services.settings_service import get_approval_thresholds
        t = get_approval_thresholds(db)
        threshold_auto = float(t["threshold_auto"] or APPROVAL_THRESHOLD_AUTO)
        threshold_manager = float(t["threshold_manager"] or APPROVAL_THRESHOLD_MANAGER)
    if level == "manager":
        return f"Importe {amount:.2f}€ entre {threshold_auto:.0f}€ y {threshold_manager:.0f}€ — requiere manager"
    return f"Importe {amount:.2f}€ ≥ {threshold_manager:.0f}€ — requiere director"


# Role capability matrix. An `admin` role can approve any level.
# Future: `manager` role approves manager; `director` role approves manager/director.
# For MVP with only employee/admin/viewer roles in DB, admin approves all non-auto.
_ROLE_CAN_APPROVE = {
    "admin": {"auto", "manager", "director"},
    "manager": {"auto", "manager"},
    "director": {"auto", "manager", "director"},
    # employee and viewer cannot approve anything above auto (and auto doesn't need approval)
}


def _can_approve(role: str, level: str) -> bool:
    """Check if a role can approve a given level."""
    if level == "auto":
        return True  # anyone could "re-confirm" an auto level but it shouldn't hit approve endpoint
    return level in _ROLE_CAN_APPROVE.get(role, set())


@router.post("/{receipt_id}/approve", response_model=ApproveRejectResult)
def approve_receipt(
    receipt_id: str,
    db: Session = Depends(get_db),
    # Auth: prefer JWT via get_current_user for audit trail; keep X-User-Role
    # as a fallback so the DEV_MODE and legacy callers keep working.
    current_user: Optional[Employee] = Depends(get_current_user_optional),
    x_user_role: str = Header(default="admin", alias="X-User-Role"),
):
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")
    if receipt.status not in ("pending", "review", "flagged"):
        raise HTTPException(status_code=409, detail=f"Cannot approve receipt with status '{receipt.status}'")

    level = receipt.approval_level or _calculate_approval_level(
        float(receipt.amount) if receipt.amount else None,
        db=db,
    )

    # Role to use for permission check: prefer authenticated user's role, fallback to header
    effective_role = current_user.role if current_user else x_user_role
    if not _can_approve(effective_role, level):
        raise HTTPException(
            status_code=403,
            detail=f"Role '{effective_role}' cannot approve level '{level}'"
        )

    has_match = db.query(Match).filter(Match.receipt_id == receipt_id).first()
    receipt.status = "matched" if has_match else "approved"
    receipt.approved_at = datetime.utcnow()
    # Audit trail: record who approved manually. Auto-approvals leave approved_by=None.
    if current_user:
        receipt.approved_by = current_user.id
    if not receipt.approval_level:
        receipt.approval_level = level
    db.commit()
    return ApproveRejectResult(status=receipt.status, message="Receipt approved")


@router.post("/{receipt_id}/reject", response_model=ApproveRejectResult)
def reject_receipt(receipt_id: str, reason: Optional[str] = None, db: Session = Depends(get_db)):
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")
    if receipt.status not in ("pending", "review", "flagged", "matched"):
        raise HTTPException(status_code=409, detail=f"Cannot reject receipt with status '{receipt.status}'")
    receipt.status = "rejected"
    if reason:
        receipt.notes = reason
    db.commit()
    return ApproveRejectResult(status="rejected", message="Receipt rejected")


@router.delete("/{receipt_id}", status_code=204)
def delete_receipt(receipt_id: str, db: Session = Depends(get_db)):
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        raise HTTPException(status_code=404, detail="Receipt not found")
    db.delete(receipt)
    db.commit()
    return Response(status_code=204)


# ── Employee Expense Excel Template ──────────────────────────────


CATEGORY_MAP_ES = {
    "transporte": "transport",
    "comidas": "meals",
    "alojamiento": "lodging",
    "material": "supplies",
    "entretenimiento": "entertainment",
    "servicios": "utilities",
    "otros": "other",
}

PAYMENT_MAP_ES = {
    "visa": "card",
    "tarjeta": "card",
    "efectivo": "cash",
    "propio": "cash",
    "transferencia": "transfer",
}


@router.get("/template/expense-excel")
def download_expense_template():
    """Generate and return a standard employee expense Excel template."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    headers = ["Fecha", "Comercio / Concepto", "Importe (EUR)", "Metodo de Pago", "Categoria", "Notas"]
    widths = [14, 30, 16, 18, 20, 30]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w

    # Example row
    example = ["15/03/2026", "Hotel Santo Mauro Madrid", 90.00, "Visa", "Alojamiento", "Cena cliente"]
    for i, val in enumerate(example, 1):
        cell = ws.cell(row=2, column=i, value=val)
        cell.font = Font(italic=True, color="94A3B8")
        cell.border = thin_border

    # Data validations for dropdown columns
    payment_dv = DataValidation(
        type="list",
        formula1='"Visa,Efectivo,Propio,Transferencia"',
        allow_blank=True,
    )
    payment_dv.error = "Selecciona: Visa, Efectivo, Propio o Transferencia"
    payment_dv.errorTitle = "Metodo no valido"
    ws.add_data_validation(payment_dv)
    payment_dv.add(f"D2:D500")

    category_dv = DataValidation(
        type="list",
        formula1='"Transporte,Comidas,Alojamiento,Material,Entretenimiento,Servicios,Otros"',
        allow_blank=True,
    )
    category_dv.error = "Selecciona una categoria valida"
    category_dv.errorTitle = "Categoria no valida"
    ws.add_data_validation(category_dv)
    category_dv.add(f"E2:E500")

    # Number format for amount column
    for row in range(2, 501):
        ws.cell(row=row, column=3).number_format = '#,##0.00'

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_gastos_expensiq.xlsx"},
    )


@router.post("/import-expense-excel")
async def import_expense_excel(
    file: UploadFile = File(...),
    employee_id: str = Form(...),
    db: Session = Depends(get_db),
):
    """Import an employee expense Excel and create Receipt records."""
    import io
    from openpyxl import load_workbook

    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel file has no active sheet")

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
    created = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or len(row) < 3:
            continue
        date_val, merchant_val, amount_val = row[0], row[1], row[2]
        payment_val = row[3] if len(row) > 3 else None
        category_val = row[4] if len(row) > 4 else None
        notes_val = row[5] if len(row) > 5 else None

        # Skip empty rows
        if not amount_val and not merchant_val:
            continue

        # Parse amount
        if isinstance(amount_val, (int, float)):
            amount = float(amount_val)
        else:
            try:
                amount = float(str(amount_val).replace(",", ".").replace(" ", ""))
            except (ValueError, TypeError):
                errors.append(f"Fila {i}: importe invalido '{amount_val}'")
                continue

        # Parse date
        receipt_date = None
        if isinstance(date_val, datetime):
            receipt_date = date_val.date()
        elif isinstance(date_val, DateType):
            receipt_date = date_val
        elif date_val:
            for dfmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                try:
                    receipt_date = datetime.strptime(str(date_val).strip(), dfmt).date()
                    break
                except ValueError:
                    continue

        # Map payment method
        pm_raw = str(payment_val or "").strip().lower()
        payment_method = PAYMENT_MAP_ES.get(pm_raw, "card")

        # Map category
        cat_raw = str(category_val or "").strip().lower()
        category = CATEGORY_MAP_ES.get(cat_raw, "other")

        level = _calculate_approval_level(amount, db=db)
        receipt = Receipt(
            id=uuid.uuid4(),
            employee_id=employee_id,
            merchant=str(merchant_val or "").strip() or None,
            date=receipt_date,
            amount=amount,
            currency="EUR",
            category=category,
            payment_method=payment_method,
            notes=str(notes_val or "").strip() or None,
            status="approved" if level == "auto" else "pending",
            approved_at=datetime.utcnow() if level == "auto" else None,
            ocr_provider="excel_import",
            ocr_confidence=1.0,
            approval_level=level,
        )
        db.add(receipt)
        created += 1

    db.commit()
    wb.close()
    logger.info("Excel import for employee %s: %d receipts created", employee_id, created)

    return {
        "created": created,
        "errors": errors[:10],
    }


def _process_receipt_ocr(receipt_id: str, file_content: bytes, filename: str):
    """Background task: run OCR on the uploaded file and update the receipt."""
    from app.db.session import SessionLocal
    from app.ocr.processor import ocr_processor
    from app.services.categorizer import ExpenseCategorizer

    db = SessionLocal()
    try:
        receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
        if not receipt:
            logger.error("Receipt %s not found for OCR processing", receipt_id)
            return

        try:
            result = ocr_processor.process(file_content, filename)
            receipt.merchant = result.get("merchant")
            receipt.date = result.get("date")
            receipt.amount = result.get("amount")
            receipt.currency = result.get("currency", "EUR")
            receipt.tax = result.get("tax")
            receipt.ocr_raw_text = result.get("raw_text")
            receipt.ocr_confidence = result.get("confidence")
            receipt.ocr_provider = result.get("provider")
            receipt.ocr_processed_at = datetime.utcnow()
            receipt.payment_method = result.get("payment_method")
            receipt.line_items = result.get("line_items")

            # Categorize
            categorizer = ExpenseCategorizer()
            receipt.category = categorizer.categorize(receipt.merchant)

            # Calculate approval level and auto-approve small amounts.
            # Bug fix Sprint 1 (C1): previously all receipts landed pending even when
            # approval_level was "auto". Now auto-tier is marked approved directly;
            # manager/director tiers remain pending and appear in the approvals queue.
            receipt.approval_level = _calculate_approval_level(
                float(receipt.amount) if receipt.amount else None,
                db=db,
            )
            if receipt.approval_level == "auto":
                receipt.status = "approved"
                receipt.approved_at = datetime.utcnow()
            else:
                receipt.status = "pending"

            logger.info("OCR processed receipt %s: %s %s %s",
                        receipt_id, receipt.merchant, receipt.amount, receipt.currency)
        except Exception as e:
            logger.error("OCR failed for receipt %s: %s", receipt_id, e)
            receipt.status = "flagged"
            receipt.notes = f"OCR error: {e}"

        db.commit()
    finally:
        db.close()
